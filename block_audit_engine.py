"""
block_audit_engine.py
Answers: "placements we've flagged to block — are they actually blocked?"

In the AdLib Insights workbook, the Site Overview and App Overview sheets carry a
'Final ... Name' column whose value is the literal sentinel "Block" when that
placement is supposed to be blocked. Any such row that still shows impressions /
billable spend is a BLOCK THAT ISN'T BEING ENFORCED — a leak.

This engine finds those leaks and attributes the wasted spend/impressions to each
Business Unit (partner), Client, Product, and Strategy.
"""
import pandas as pd
import numpy as np
import re
import datetime
from openpyxl import load_workbook

SENTINELS = {"block", "blocked"}

# --- Heuristic auto-block: gaming + junk + unresolved bundle apps -----------
_GAMING = re.compile(
    r"\b(puzzle|casino|slots?|bingo|solitaire|mahjong|bubble\s?(shoot|pop)|"
    r"block\s?(puzzle|blast|hexa|mania|party|craft|jam)|match\s?3|arcade|tycoon|clash|"
    r"sudoku|jackpot|poker|hexa|jewel\s?(blast|quest)|candy\s?crush|saga|idle\s|"
    r"trivia\s?crack|racing\s?game|ludo|tetris|2048|gacha|tower\s?defen|zombie|"
    r"battle\s?(royale|craft|land)|shooter|word\s?(trip|connect|search|cross|calm|link)|"
    r"gems?\s?(blast|crush)|gold\s?miner|dragon\s?(city|mania|blast)|merge\s?(dragons|"
    r"mansion|magic)|\.io\b|coin\s?master|spin\s?to\s?win)\b", re.I)
_JUNK = re.compile(
    r"\b(photo\s?edit|beauty\s?cam|selfie|flashlight|battery\s?(saver|doctor)|cleaner|"
    r"booster|antivirus|qr\s?(scanner|code)|wallpaper|ringtone|keyboard|file\s?manager|"
    r"compass|magnifier|clean\s?master|du\s?battery)\b", re.I)
_BUNDLE = re.compile(r"^((com|net|org|io|app)\.[a-z0-9_.]+|\d{6,})$", re.I)


def classify_app_junk(name):
    """Return (category, reason) if an app should be auto-blocked, else (None, None)."""
    if not isinstance(name, str) or not name.strip():
        return (None, None)
    n = name.strip()
    if _BUNDLE.match(n):
        return ("Unresolved bundle", "unidentifiable app (raw bundle/ID)")
    if _GAMING.search(n):
        return ("Gaming app", "gaming inventory (block-by-default)")
    if _JUNK.search(n):
        return ("Junk app", "low-value utility/photo/junk app")
    return (None, None)

# Column-name token groups we need to pull when streaming the big sheets.
NEED_TOKENS = [("final",), ("site", "domain"), ("app", "name"), ("app", "id"),
               ("bundle",), ("impression",), ("click",), ("conversion",), ("conv",),
               ("billable", "spend"), ("spend",), ("cost",), ("date",),
               ("business", "unit"), ("client",), ("product",), ("strategy",),
               ("campaign",)]

SHEET_CONFIG = {
    "Site Overview": {"kind": "site"},
    "App Overview": {"kind": "app"},
}

# Which master-blocklist tab covers which product. "All Products" covers everything.
_PRODUCT_TAB = {"CTV": "CTV", "Online Audio": "Audio", "Social Mirror": "Social Mirror"}
_BLANK_DATE = datetime.date(1970, 1, 1)  # blank date_added = blocked before any dated row


def _is_blocked_for_product(product, entry):
    """True if this placement's blocklist membership covers this product."""
    tabs = entry.get("tabs", set())
    if "All Products" in tabs:
        return True
    tab = _PRODUCT_TAB.get(product)
    return bool(tab and tab in tabs)


def _blocked_since(product, entry):
    """Earliest date this product is blocked for this placement, or None. A blank
    date on an applicable list counts as 'blocked long ago'."""
    td = entry.get("tab_dates", {})
    tabs = entry.get("tabs", set())
    cands = []
    if "All Products" in tabs:
        cands.append(td.get("All Products") or _BLANK_DATE)
    tab = _PRODUCT_TAB.get(product)
    if tab and tab in tabs:
        cands.append(td.get(tab) or _BLANK_DATE)
    return min(cands) if cands else None


def _pick_client(cols):
    """The 'Client' column, not 'Client Business Unit'."""
    low = {str(c).lower().strip(): c for c in cols}
    if "client" in low:
        return low["client"]
    for lc, orig in low.items():
        if "client" in lc and "business" not in lc and "unit" not in lc:
            return orig
    return None


def _find(cols, *tokens_any, exclude=()):
    """Return first column whose lowercased name contains ALL tokens in any group,
    skipping any column that contains an `exclude` token (e.g. 'pool' so that
    'Campaign ID' matches but 'Campaign Pool ID' doesn't)."""
    low = {c.lower(): c for c in cols}
    for group in tokens_any:
        for lc, orig in low.items():
            if any(x in lc for x in exclude):
                continue
            if all(t in lc for t in group):
                return orig
    return None


def _detect(df, kind):
    cols = df.columns
    final = _find(cols, ("final", "site"), ("final", "app"), ("final", "domain"))
    raw = _find(cols, ("site", "domain"), ("app", "name")) or final
    return {
        "final": final,
        "raw": raw,
        "impr": _find(cols, ("impression",)),
        "clicks": _find(cols, ("click",)),
        "spend": _find(cols, ("billable", "spend"), ("spend",), ("cost",)),
        "conv": _find(cols, ("click", "conversion"), ("conversion",), ("conv",)),
        "date": _find(cols, ("date",)),
        "bu": _find(cols, ("business", "unit")),
        "client": _pick_client(cols),
        "product": _find(cols, ("product",)),
        "strategy": _find(cols, ("strategy", "type"), ("strategy",)),
        "strategy_name": _find(cols, ("strategy", "name")),
        "app_id": _find(cols, ("app", "id"), ("app", "bundle"), ("bundle",)),
        "campaign": _find(cols, ("campaign", "name"), exclude=("pool",)),
        "campaign_id": _find(cols, ("campaign", "id"), ("campaign", "#"), ("campaign", "number"), exclude=("pool",)),
    }


def _normalize(df, c):
    out = pd.DataFrame()
    out["placement"] = df[c["raw"]].astype(str)
    out["final"] = df[c["final"]].astype(str).str.strip()
    out["impressions"] = pd.to_numeric(df[c["impr"]], errors="coerce").fillna(0) if c["impr"] else 0
    out["clicks"] = pd.to_numeric(df[c["clicks"]], errors="coerce").fillna(0) if c["clicks"] else 0
    out["spend"] = pd.to_numeric(df[c["spend"]], errors="coerce").fillna(0) if c["spend"] else 0
    conv_cols = [col for col in df.columns if "conversion" in str(col).lower()]
    if conv_cols:
        out["conversions"] = sum(pd.to_numeric(df[col], errors="coerce").fillna(0) for col in conv_cols)
        out["_has_conv"] = True
    else:
        out["conversions"] = 0
        out["_has_conv"] = False
    out["served_date"] = pd.to_datetime(df[c["date"]], errors="coerce") if c["date"] else pd.NaT
    out["app_id"] = df[c["app_id"]].astype(str) if c.get("app_id") else out["placement"]
    # Display name: real app name, or fall back to App ID when the name is NA/blank
    # (so unresolved apps stay distinct by ID instead of collapsing into one "NA").
    _bad = out["placement"].str.strip().str.lower().isin({"na", "nan", "none", "", "(not set)"})
    out["disp"] = out["placement"].where(~_bad, out["app_id"])
    for dim in ("bu", "client", "product", "strategy"):
        out[dim] = (df[c[dim]].astype(str) if c[dim] else "(not in export)")
    out["strategy_name"] = df[c["strategy_name"]].astype(str) if c.get("strategy_name") else out["strategy"]
    out["campaign"] = df[c["campaign"]].astype(str) if c.get("campaign") else "(not in export)"
    # IDs come out of Excel as floats ("12345.0") — normalize to clean strings.
    out["campaign_id"] = (df[c["campaign_id"]].astype(str).str.strip()
                          .str.replace(r"\.0$", "", regex=True)
                          if c.get("campaign_id") else "")
    out["is_block"] = out["final"].str.lower().isin(SENTINELS)
    out["is_unresolved"] = df[c["final"]].isna() | (out["final"].str.lower().isin({"nan", ""}))
    return out


def _rollup(leak, dim):
    cols = [dim, "leaked_impressions", "leaked_clicks", "ctr", "leaked_conversions", "leaked_spend", "placements"]
    if leak.empty:
        return pd.DataFrame(columns=cols)
    g = (leak.groupby(dim)
         .agg(leaked_impressions=("impressions", "sum"), leaked_clicks=("clicks", "sum"),
              leaked_conversions=("conversions", "sum"), leaked_spend=("spend", "sum"),
              placements=("placement", "nunique"))
         .reset_index())
    g["ctr"] = np.where(g["leaked_impressions"] > 0, g["leaked_clicks"] / g["leaked_impressions"], 0)
    return g.sort_values("ctr", ascending=False)


# Products where a LOW CTR is expected (they don't drive clicks), so they're
# excluded from the low-CTR site check. Includes the "Audio" alias.
LOW_CTR_EXCLUDED_PRODUCTS = {"CTV", "Social Mirror CTV", "Online Audio", "Audio"}


def low_ctr_sites_by_client(allp, min_impr=5000, ctr_multiple=3.0,
                            ctr_floor=0.0005, conv_rate_keep=0.0003):
    """Per-client, per-site watchlist of SITES with abnormally LOW CTR.

    A site is flagged only when it is BOTH far below its product's own CTR norm
    (<= 1/ctr_multiple of the pooled product CTR) AND under an absolute ctr_floor
    (0.05% by default). CTV / Social Mirror CTV / Online Audio are excluded — a low
    CTR is expected on those non-click products. Sites converting efficiently
    (conv/impr >= conv_rate_keep) are never flagged, even at low CTR — they're
    working. Returns one row per client + site (+ product, since CTR norms are
    product-specific), sorted by wasted spend.
    """
    if allp is None or not len(allp):
        return pd.DataFrame()
    site = allp[(allp["placement_type"] == "site") & (allp["impressions"] > 0)].copy()
    site["product"] = site["product"].astype(str).str.strip()
    site = site[~site["product"].isin(LOW_CTR_EXCLUDED_PRODUCTS)]
    site = site[~site["product"].str.lower().isin({"", "nan", "none", "(not in export)"})]
    if site.empty:
        return pd.DataFrame()

    g = (site.groupby(["bu", "client", "product", "placement"], dropna=False)
         .agg(impressions=("impressions", "sum"), clicks=("clicks", "sum"),
              conversions=("conversions", "sum"), spend=("spend", "sum"))
         .reset_index()
         .rename(columns={"bu": "business_unit", "client": "Client", "placement": "site"}))
    g["ctr"] = np.where(g["impressions"] > 0, g["clicks"] / g["impressions"], 0)

    # Per-product norm = pooled CTR across all sites/clients on that product.
    norms = (g.groupby("product").apply(
        lambda x: (x["clicks"].sum() / x["impressions"].sum()) if x["impressions"].sum() else 0)
        .rename("product_ctr").reset_index())
    g = g.merge(norms, on="product", how="left")
    g["x_over_norm"] = np.where(g["product_ctr"] > 0, g["ctr"] / g["product_ctr"], 0)
    g["conv_rate"] = np.where(g["impressions"] > 0, g["conversions"] / g["impressions"], 0)

    material = g["impressions"] >= min_impr
    below_norm = (g["product_ctr"] > 0) & (g["x_over_norm"] <= (1.0 / ctr_multiple))
    below_floor = g["ctr"] <= ctr_floor
    converting = g["conv_rate"] >= conv_rate_keep
    flagged = g[material & below_norm & below_floor & ~converting].copy()
    if flagged.empty:
        return flagged
    flagged["reason"] = flagged.apply(
        lambda r: f"CTR {r['ctr']*100:.3f}% is {r['x_over_norm']:.2f}× the {r['product']} "
                  f"norm ({r['product_ctr']*100:.3f}%) and below the {ctr_floor*100:.2f}% floor",
        axis=1)
    return flagged.sort_values("spend", ascending=False).reset_index(drop=True)


def auto_site_blocks(allp, min_impr=10000, ctr_floor=0.0005, conv_rate_keep=0.0003,
                     exclude_keys=None):
    """Sites to auto-add to the recommended block list. Judged ACROSS THE BOARD
    (pooled over every client/campaign, not per client): a site delivering
    meaningful volume on click-driven products with a very low CTR (<= ctr_floor)
    AND little-to-no conversions (conv/impr < conv_rate_keep) is dead weight for
    everyone, so it's safe to block globally. CTV / Social Mirror CTV / Online
    Audio are excluded — low CTR is expected there. Returns the same schema as
    auto_app_blocks so it merges straight into the site block list.
    """
    cols = ["name", "app_id", "products", "impressions", "clicks", "ctr", "spend",
            "conversions", "category", "reason"]
    if allp is None or not len(allp):
        return pd.DataFrame(columns=cols)
    site = allp[(allp["placement_type"] == "site") & (allp["impressions"] > 0)].copy()
    site["product"] = site["product"].astype(str).str.strip()
    site = site[~site["product"].isin(LOW_CTR_EXCLUDED_PRODUCTS)]
    site = site[~site["product"].str.lower().isin({"", "nan", "none", "(not in export)"})]
    if site.empty:
        return pd.DataFrame(columns=cols)
    g = (site.groupby("placement")
         .agg(products=("product", lambda s: ", ".join(sorted({str(p) for p in s if str(p).strip()}))),
              impressions=("impressions", "sum"), clicks=("clicks", "sum"),
              conversions=("conversions", "sum"), spend=("spend", "sum"))
         .reset_index().rename(columns={"placement": "name"}))
    g = g[~g["name"].astype(str).str.strip().str.lower().isin({"na", "nan", "none", ""})]
    # Don't re-recommend a site already flagged "Block" or already on the master
    # blocklist — those are handled; we only want NEW dead sites.
    if exclude_keys:
        g = g[~g["name"].astype(str).str.strip().str.lower().isin(set(exclude_keys))]
    g["ctr"] = np.where(g["impressions"] > 0, g["clicks"] / g["impressions"], 0)
    g["conv_rate"] = np.where(g["impressions"] > 0, g["conversions"] / g["impressions"], 0)
    flag = ((g["impressions"] >= min_impr) & (g["ctr"] <= ctr_floor) &
            (g["conv_rate"] < conv_rate_keep))
    g = g[flag].copy()
    if g.empty:
        return pd.DataFrame(columns=cols)
    g["app_id"] = g["name"]
    g["category"] = "Low CTR / no conv"
    g["reason"] = g.apply(
        lambda r: f"CTR {r['ctr']*100:.3f}% with {int(round(r['conversions']))} conversions "
                  f"across {int(r['impressions']):,} impressions (all clients)", axis=1)
    return g[cols].sort_values("spend", ascending=False).reset_index(drop=True)


def auto_high_ctr_site_blocks(allp, min_impr=10000, ctr_multiple=3.0, ctr_floor=0.01,
                              exclude_keys=None):
    """Sites to auto-add to the block list for ABNORMALLY HIGH CTR. Pooled across all
    clients, a site whose CTR is both >= ctr_floor (absolute) AND >= ctr_multiple× its
    product's own norm is a likely invalid-traffic / bot signal worth blocking. CTV /
    Social Mirror CTV / Online Audio are excluded — their near-zero click norms make
    the ratio meaningless. Same schema as auto_app_blocks so it merges into the list.
    """
    cols = ["name", "app_id", "products", "impressions", "clicks", "ctr", "spend",
            "conversions", "category", "reason"]
    if allp is None or not len(allp):
        return pd.DataFrame(columns=cols)
    site = allp[(allp["placement_type"] == "site") & (allp["impressions"] > 0)].copy()
    site["product"] = site["product"].astype(str).str.strip()
    site = site[~site["product"].isin(LOW_CTR_EXCLUDED_PRODUCTS)]
    site = site[~site["product"].str.lower().isin({"", "nan", "none", "(not in export)"})]
    if site.empty:
        return pd.DataFrame(columns=cols)
    g = (site.groupby(["placement", "product"])
         .agg(impressions=("impressions", "sum"), clicks=("clicks", "sum"),
              conversions=("conversions", "sum"), spend=("spend", "sum"))
         .reset_index())
    g["ctr"] = np.where(g["impressions"] > 0, g["clicks"] / g["impressions"], 0)
    norms = (g.groupby("product").apply(
        lambda x: (x["clicks"].sum() / x["impressions"].sum()) if x["impressions"].sum() else 0)
        .rename("product_ctr").reset_index())
    g = g.merge(norms, on="product", how="left")
    g["x_over_norm"] = np.where(g["product_ctr"] > 0, g["ctr"] / g["product_ctr"], 0)
    flag = ((g["impressions"] >= min_impr) & (g["ctr"] >= ctr_floor) &
            (g["product_ctr"] > 0) & (g["x_over_norm"] >= ctr_multiple))
    g = g[flag].copy()
    g = g[~g["placement"].astype(str).str.strip().str.lower().isin({"na", "nan", "none", ""})]
    if exclude_keys:
        g = g[~g["placement"].astype(str).str.strip().str.lower().isin(set(exclude_keys))]
    if g.empty:
        return pd.DataFrame(columns=cols)
    # One row per site — keep the product on which it's most extreme.
    g = g.sort_values("x_over_norm", ascending=False).drop_duplicates("placement", keep="first")
    g = g.rename(columns={"placement": "name", "product": "products"})
    g["app_id"] = g["name"]
    g["category"] = "High CTR"
    g["reason"] = g.apply(
        lambda r: f"CTR {r['ctr']*100:.2f}% is {r['x_over_norm']:.1f}× the {r['products']} "
                  f"norm — abnormally HIGH (possible invalid traffic)", axis=1)
    return g[cols].sort_values("spend", ascending=False).reset_index(drop=True)


def _stream_sheet(ws):
    """Stream a worksheet in read_only mode, pulling only the columns we need.
    Keeps peak memory low (never materializes the full sheet) and avoids loading
    the whole 40MB workbook. Returns (slim_df, row_count)."""
    it = ws.iter_rows(values_only=True)
    try:
        header = next(it)
    except StopIteration:
        return None, 0
    wanted = {}
    for i, h in enumerate(header):
        if h is None:
            continue
        hl = str(h).lower()
        if any(all(t in hl for t in grp) for grp in NEED_TOKENS):
            wanted[str(h)] = i
    if not wanted:
        return None, 0
    data = {name: [] for name in wanted}
    n = 0
    for r in it:
        for name, i in wanted.items():
            data[name].append(r[i] if i < len(r) else None)
        n += 1
    return pd.DataFrame(data), n


def audit_block_leak(path_or_buffer=None, blocklist=None, frames=None):
    truncation = {}
    norm_frames = []
    if frames is not None:
        # Frames handed in directly (automated pull) — no xlsx read, dtypes kept.
        for kind, df in (("site", frames.get("site")), ("app", frames.get("app"))):
            if df is None or not len(df):
                continue
            c = _detect(df, kind)
            if not c["final"]:
                continue
            norm = _normalize(df, c)
            norm["placement_type"] = kind
            norm_frames.append(norm)
            truncation[f"{kind.title()} Overview"] = False
    else:
        wb = load_workbook(path_or_buffer, read_only=True, data_only=True)
        try:
            for sheet, cfg in SHEET_CONFIG.items():
                if sheet not in wb.sheetnames:
                    continue
                df, nrows = _stream_sheet(wb[sheet])
                if df is None:
                    continue
                c = _detect(df, cfg["kind"])
                if not c["final"]:
                    continue
                norm = _normalize(df, c)
                norm["placement_type"] = cfg["kind"]
                norm_frames.append(norm)
                truncation[sheet] = nrows >= 100000  # Tap export row cap heuristic
                df = None
        finally:
            wb.close()

    if not norm_frames:
        raise ValueError("No Site/App Overview data with a 'Final ... Name' column found.")

    allp = pd.concat(norm_frames, ignore_index=True)
    norm_frames.clear()
    import gc
    gc.collect()
    leak = allp[allp["is_block"] & (allp["impressions"] > 0)]
    unresolved = allp[allp["is_unresolved"] & (allp["impressions"] > 0)]

    # Recency: a trailing report shows impressions that may predate the block.
    # Where a date exists (apps), use last-served to tell "stopped mid-window"
    # (block likely took hold) from "still serving at the window edge" (verify).
    window_end = allp["served_date"].max()
    window_start = allp["served_date"].min()
    ACTIVE_DAYS = 1  # served on/after window_end - 1 day = still active

    by_type = (leak.groupby("placement_type")
               .agg(rows=("placement", "size"), placements=("placement", "nunique"),
                    impressions=("impressions", "sum"), spend=("spend", "sum")).reset_index())

    offenders = (leak.groupby(["placement", "placement_type"])
                 .agg(impressions=("impressions", "sum"), clicks=("clicks", "sum"),
                      conversions=("conversions", "sum"), spend=("spend", "sum"),
                      last_served=("served_date", "max"))
                 .reset_index().sort_values("spend", ascending=False))
    offenders["ctr"] = np.where(offenders["impressions"] > 0, offenders["clicks"] / offenders["impressions"], 0)
    if pd.notna(window_end):
        no_date = offenders["last_served"].isna()
        offenders["days_since_last"] = (window_end - offenders["last_served"]).dt.days
        offenders["still_active"] = (offenders["days_since_last"] <= ACTIVE_DAYS).astype("object")
        offenders.loc[no_date, "still_active"] = pd.NA
    else:
        offenders["days_since_last"] = pd.NA
        offenders["still_active"] = pd.NA
    offenders["last_served"] = offenders["last_served"].dt.strftime("%Y-%m-%d").fillna("(no date)")

    active = offenders[offenders["still_active"] == True]  # noqa: E712

    # Distinct names already flagged "Block" (for copyable AdLib filter syntax)
    blocked = allp[allp["is_block"]]
    block_names = {
        "site": sorted(blocked.loc[blocked["placement_type"] == "site", "placement"].dropna().unique().tolist()),
        "app": sorted(blocked.loc[blocked["placement_type"] == "app", "placement"].dropna().unique().tolist()),
    }

    # Candidates for the AI pass: placements NOT already flagged Block, with real
    # delivery. Aggregate distinct by spend, list all products each ran on, and
    # (apps) carry the App ID used for the AdLib block.
    already = {k: set(v) for k, v in block_names.items()}
    cand = allp[(~allp["is_block"]) & (allp["impressions"] > 0)]

    VALID_PRODUCTS = {"Display", "Social Mirror", "Video", "CTV", "Native Display",
                      "Native Video", "Social Mirror CTV", "Online Audio", "Audio"}

    def _products(series):
        vals = [p for p in series.dropna().unique().tolist() if p in VALID_PRODUCTS]
        return ", ".join(sorted(set(vals))) if vals else ""

    def _candidates(kind):
        sub = cand[cand["placement_type"] == kind]
        if sub.empty:
            return pd.DataFrame(columns=["name", "app_id", "products", "impressions", "clicks", "spend"])
        d = (sub.groupby("placement")
             .agg(app_id=("app_id", "first"), products=("product", _products),
                  impressions=("impressions", "sum"), clicks=("clicks", "sum"),
                  conversions=("conversions", "sum"), spend=("spend", "sum"))
             .reset_index().rename(columns={"placement": "name"}))
        d = d[~d["name"].isin(already.get(kind, set()))]
        # Also drop anything already covered by the external blocklist (what you've
        # pushed). AdLib enforcement lags in the data, so a freshly-blocked site can
        # still be serving (~is_block) — without this it would keep getting
        # recommended. Product-aware: only drop if blocked for every product it ran on.
        if blocklist:
            def _covered(row):
                mk = (str(row["app_id"]).strip().lower() if kind == "app"
                      else str(row["name"]).strip().lower())
                entry = blocklist.get(mk)
                if not entry:
                    return False
                prods = [p.strip() for p in str(row["products"]).split(",") if p.strip()]
                if not prods:
                    return True  # on the list, no product detail -> treat as covered
                return all(_is_blocked_for_product(p, entry) for p in prods)
            if len(d):
                d = d[~d.apply(_covered, axis=1)]
        return d.sort_values("spend", ascending=False).reset_index(drop=True)

    candidates = {"site": _candidates("site"), "app": _candidates("app")}

    # Deterministic auto-block: gaming + junk + unresolved bundle apps (every run).
    app_c = candidates["app"]
    auto_rows = []
    for _, r in app_c.iterrows():
        cat, reason = classify_app_junk(r["name"])
        if cat:
            row = r.to_dict()
            row["category"] = cat
            row["reason"] = reason
            row["ctr"] = (row["clicks"] / row["impressions"]) if row["impressions"] else 0
            auto_rows.append(row)
    auto_app_blocks = pd.DataFrame(auto_rows, columns=[
        "name", "app_id", "products", "impressions", "clicks", "ctr", "spend", "category", "reason"])

    # Combined placements grid: ALL delivery (blocked + non-blocked). Apps use the
    # display name (App ID when the name is NA). Carries last-served + blocked flag.
    alld = allp[allp["impressions"] > 0].copy()
    alld["disp_name"] = alld["disp"].where(alld["placement_type"] == "app", alld["placement"])
    alld = alld[~alld["disp_name"].str.strip().str.lower().isin({"na", "nan", "none", ""})]
    topbase = (alld.groupby(["disp_name", "placement_type"])
               .agg(products=("product", _products), impressions=("impressions", "sum"),
                    clicks=("clicks", "sum"), conversions=("conversions", "sum"),
                    spend=("spend", "sum"), last_dt=("served_date", "max"),
                    blocked=("is_block", "max"))
               .reset_index().rename(columns={"disp_name": "name"}))
    topbase["ctr"] = np.where(topbase["impressions"] > 0, topbase["clicks"] / topbase["impressions"], 0)
    topbase["cpm"] = np.where(topbase["impressions"] > 0, topbase["spend"] / topbase["impressions"] * 1000, 0)
    # Serving status: still serving at the window edge (within ACTIVE_DAYS of the
    # last date in the file), stopped, or unknown (no date).
    if pd.notna(window_end):
        days = (window_end - topbase["last_dt"]).dt.days
        topbase["serving"] = np.where(topbase["last_dt"].isna(), "—",
                                      np.where(days <= ACTIVE_DAYS, "serving", "stopped"))
    else:
        topbase["serving"] = "—"
    topbase["last_served"] = topbase["last_dt"].dt.strftime("%Y-%m-%d").fillna("—")
    topbase = topbase.drop(columns=["last_dt"])
    # Share of ALL delivery impressions (denominator is the full dataset, not
    # just the rows that survive the head() cut below).
    _tot_impr = float(alld["impressions"].sum())
    topbase["pct_impr"] = np.where(_tot_impr > 0, topbase["impressions"] / _tot_impr, 0)
    top_placements = topbase.sort_values("impressions", ascending=False).head(300)

    # Per-placement-per-product delivery, so the app can compute Block-impact against
    # the RECOMMENDED block set (which is only known after the AI step).
    VALID = {"Display", "Social Mirror", "Video", "CTV", "Native Display",
             "Native Video", "Social Mirror CTV", "Online Audio", "Audio"}
    imp_df = allp[allp["product"].isin(VALID) & (allp["impressions"] > 0)].copy()
    imp_df["match_key"] = np.where(imp_df["placement_type"] == "app",
                                   imp_df["app_id"].astype(str).str.strip().str.lower(),
                                   imp_df["placement"].astype(str).str.strip().str.lower())
    delivery_pp = (imp_df.groupby(["match_key", "placement_type", "product"])
                   .agg(impressions=("impressions", "sum"), spend=("spend", "sum"))
                   .reset_index())

    # Placement-level rows by grain, so the app can recompute Partner/Client/Strategy
    # CTR after removing recommended-block placements ("impact of recommendations").
    _wl = allp[allp["impressions"] > 0].copy()
    _wl["match_key"] = np.where(_wl["placement_type"] == "app",
                                _wl["app_id"].astype(str).str.strip().str.lower(),
                                _wl["placement"].astype(str).str.strip().str.lower())
    wl_src = (_wl.groupby(["match_key", "bu", "client", "product", "strategy_name"], dropna=False)
              .agg(impressions=("impressions", "sum"), clicks=("clicks", "sum"))
              .reset_index())

    # Per-placement-per-strategy-type delivery, for "Block impact by strategy".
    delivery_strat = (_wl.groupby(["match_key", "strategy"], dropna=False)
                      .agg(impressions=("impressions", "sum"), spend=("spend", "sum"))
                      .reset_index())

    has_conv = bool(allp["_has_conv"].any())

    # Low-CTR site watchlist: sites (by client) that are BOTH far below their
    # product's CTR norm AND under an absolute floor. Excludes CTV/SM CTV/Audio.
    low_ctr_sites = low_ctr_sites_by_client(allp)

    # Sites to auto-add to the recommended block list: across ALL clients, low CTR
    # with little-to-no conversions (dead weight for everyone). Skip sites already
    # flagged "Block" or already on the master blocklist.
    _site_block_keys = {str(n).strip().lower() for n in already.get("site", set())}
    if blocklist:
        _site_block_keys |= set(blocklist.keys())
    auto_site_blocks_df = auto_site_blocks(allp, exclude_keys=_site_block_keys)
    # Abnormally HIGH CTR sites (possible invalid traffic) — also auto-block candidates.
    auto_high_ctr_site_blocks_df = auto_high_ctr_site_blocks(allp, exclude_keys=_site_block_keys)

    # Master-blocklist check: match delivery against the external blocklist sheet
    # (by domain for sites, App ID for apps) and flag anything still serving AFTER
    # its "date added" — but only on a product it's actually blocked on. A CTV-list
    # placement that served on Display is NOT a leak (it was never blocked there).
    blocklist_check = None
    blocklist_by_bu = None
    blocked_site_clients = None
    if blocklist:
        a2 = allp[allp["impressions"] > 0].copy()
        a2["match_key"] = np.where(a2["placement_type"] == "app",
                                   a2["app_id"].astype(str).str.strip().str.lower(),
                                   a2["placement"].astype(str).str.strip().str.lower())
        a2 = a2[a2["match_key"].isin(blocklist.keys())].copy()
        if len(a2):
            a2["blocked_date"] = a2.apply(
                lambda r: _blocked_since(r["product"], blocklist[r["match_key"]]), axis=1)
            bd = pd.to_datetime(a2["blocked_date"], errors="coerce")
            sd = pd.to_datetime(a2["served_date"], errors="coerce")
            a2["after_block"] = bd.notna() & sd.notna() & (sd > bd)
            a2["post_impr"] = np.where(a2["after_block"], a2["impressions"], 0)
            a2["post_spend"] = np.where(a2["after_block"], a2["spend"], 0)
            a2["display_name"] = np.where(a2["placement_type"] == "app",
                                          a2["disp"].astype(str), a2["placement"].astype(str))
            g = (a2.groupby("match_key")
                 .agg(name=("display_name", "first"), placement_type=("placement_type", "first"),
                      products=("product", _products), impressions=("impressions", "sum"),
                      spend=("spend", "sum"), post_impr=("post_impr", "sum"),
                      post_spend=("post_spend", "sum"), last_dt=("served_date", "max"))
                 .reset_index())

            # Separate grid: any CLIENT serving on blocklisted placements — so you can
            # verify that client's block settings. One row per partner+client, with the
            # blocked sites they ran, the products, and delivery metrics (incl. the
            # post-block "leak" portion).
            _BAD_CLIENT = {"nan", "none", "", "(not in export)"}

            def _site_list(series):
                # Full, de-duped, sorted list of blocklisted placements (no truncation)
                # — the dashboard shows it in a collapsible full-width drawer.
                return sorted({str(s).strip() for s in series if str(s).strip()})

            # Broken out by partner + client + strategy. AdLib's "Strategy Name" is
            # what the org calls the campaign name, and its Campaign ID is surfaced
            # as "Strategy ID". When the export carries neither, both collapse to
            # placeholders and the grain is effectively one row per client (as before).
            # Per-type unique counts for the drawer label ("3 sites · 2 apps").
            # .where() leaves NaN for the other type, which nunique ignores.
            a2["_site_key"] = a2["match_key"].where(a2["placement_type"] == "site")
            a2["_app_key"] = a2["match_key"].where(a2["placement_type"] == "app")
            cob = (a2.groupby(["bu", "client", "strategy_name", "campaign_id"], dropna=False)
                   .agg(products=("product", _products), impressions=("impressions", "sum"),
                        clicks=("clicks", "sum"), spend=("spend", "sum"),
                        post_impr=("post_impr", "sum"), post_spend=("post_spend", "sum"),
                        n_sites=("match_key", "nunique"),
                        n_site=("_site_key", "nunique"), n_app=("_app_key", "nunique"),
                        sites_list=("display_name", _site_list))
                   .reset_index().rename(columns={"bu": "business_unit", "client": "Client",
                                                  "strategy_name": "Strategy Name",
                                                  "campaign_id": "Strategy ID"}))
            cob = cob[~cob["Client"].astype(str).str.strip().str.lower().isin(_BAD_CLIENT)]
            cob["ctr"] = np.where(cob["impressions"] > 0, cob["clicks"] / cob["impressions"], 0)
            cob["sites"] = cob["sites_list"].apply(lambda lst: ", ".join(lst))
            cob = cob[["business_unit", "Client", "Strategy Name", "Strategy ID", "products",
                       "sites", "sites_list", "n_sites", "n_site", "n_app",
                       "impressions", "clicks", "ctr",
                       "spend", "post_impr", "post_spend"]]
            blocked_site_clients = cob.sort_values(
                ["post_spend", "spend"], ascending=False).reset_index(drop=True)
            g["lists"] = g["match_key"].map(
                lambda k: ", ".join(sorted(blocklist[k].get("tabs", set()))))
            g["date_added"] = g["match_key"].map(lambda k: blocklist[k].get("date_added"))
            g["leaking"] = g["post_impr"] > 0
            g["last_served"] = pd.to_datetime(g["last_dt"], errors="coerce").dt.strftime("%Y-%m-%d").fillna("—")
            g["date_added_str"] = g["date_added"].map(lambda d: d.strftime("%Y-%m-%d") if d else "—")
            g = g.drop(columns=["last_dt"])
            blocklist_check = {
                "matched": int(len(g)),
                "leaking_count": int(g["leaking"].sum()),
                "leaking_spend": float(g.loc[g["leaking"], "post_spend"].sum()),
                "rows": g.sort_values(["leaking", "post_spend"], ascending=[False, False]),
            }
            # Per-partner exposure to blocklisted inventory (product-aware): serving
            # on placements that are on the blocklist for that product.
            exp = a2[a2["blocked_date"].notna()].copy()
            if len(exp) and "bu" in exp.columns:
                blocklist_by_bu = (exp.groupby("bu")
                                   .agg(blocked_impr=("impressions", "sum"),
                                        blocked_placements=("match_key", "nunique"))
                                   .reset_index())

    summary = {
        "leaked_spend": float(leak["spend"].sum()),
        "leaked_impressions": float(leak["impressions"].sum()),
        "leaked_placements": int(leak["placement"].nunique()),
        "leaked_rows": int(len(leak)),
        "by_type": by_type.to_dict("records"),
        "truncated_sheets": [s for s, t in truncation.items() if t],
        "unresolved_placements": int(unresolved["placement"].nunique()),
        "unresolved_spend": float(unresolved["spend"].sum()),
        "window_start": window_start.strftime("%b %d") if pd.notna(window_start) else None,
        "window_end": window_end.strftime("%b %d") if pd.notna(window_end) else None,
        "active_placements": int(len(active)),
        "active_spend": float(active["spend"].sum()) if len(active) else 0.0,
        "has_dates": bool(pd.notna(window_end)),
    }

    return {
        "summary": summary,
        "offenders": offenders,
        "leak_by_bu": _rollup(leak, "bu"),
        "leak_by_client": _rollup(leak, "client"),
        "leak_by_product": _rollup(leak, "product"),
        "leak_by_strategy": _rollup(leak, "strategy"),
        "block_names": block_names,
        "candidates": candidates,
        "auto_app_blocks": auto_app_blocks,
        "auto_site_blocks": auto_site_blocks_df,
        "auto_high_ctr_site_blocks": auto_high_ctr_site_blocks_df,
        "top_placements": top_placements,
        "delivery_pp": delivery_pp,
        "delivery_strat": delivery_strat,
        "wl_src": wl_src,
        "low_ctr_sites": low_ctr_sites,
        "blocklist_check": blocklist_check,
        "blocked_site_clients": blocked_site_clients,
        "blocklist_by_bu": blocklist_by_bu,
        "has_conv": has_conv,
        "has_app_id": bool(app_c["app_id"].ne(app_c["name"]).any()) if len(app_c) else False,
    }
