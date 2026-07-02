"""
exchange_engine.py
Reads the "Exchanges Overview" sheet and flags exchanges behaving abnormally.

NOTE: the exchange grain has no conversions column, so "low conversions" can't be
measured here. What we CAN flag from this sheet:
  - CTR far above the book average at meaningful spend  -> possible invalid / bot
    traffic or a junk supply path (abnormal clicks)
  - CTR far below the book average at meaningful spend  -> likely dead/wasted or
    audio/CTV inventory not driving engagement
  - single exchanges concentrating a large share of spend
Uses openpyxl read_only streaming to stay memory-light.
"""
import pandas as pd
import numpy as np
from openpyxl import load_workbook

SHEET = "Exchanges Overview"
NEED = [("exchange",), ("client business unit",), ("business unit",), ("client",),
        ("product",), ("strategy", "type"), ("deal",), ("conversion",), ("conv",),
        ("impression",), ("click",), ("billable", "spend"), ("spend",), ("cost",)]


def _find(cols, *groups):
    low = {c.lower(): c for c in cols}
    for g in groups:
        for lc, orig in low.items():
            if all(t in lc for t in g):
                return orig
    return None


def _stream(ws):
    it = ws.iter_rows(values_only=True)
    try:
        header = next(it)
    except StopIteration:
        return None
    wanted = {}
    for i, h in enumerate(header):
        if h is None:
            continue
        hl = str(h).lower()
        if any(all(t in hl for t in g) for g in NEED):
            wanted[str(h)] = i
    if not wanted:
        return None
    data = {n: [] for n in wanted}
    for r in it:
        for n, i in wanted.items():
            data[n].append(r[i] if i < len(r) else None)
    return pd.DataFrame(data)


def analyze_exchanges(path_or_buffer, min_spend=150.0, min_impr=30000, ctr_multiple=3.0):
    wb = load_workbook(path_or_buffer, read_only=True, data_only=True)
    try:
        if SHEET not in wb.sheetnames:
            return None
        df = _stream(wb[SHEET])
    finally:
        wb.close()
    if df is None or df.empty:
        return None

    ex = _find(df.columns, ("exchange",))
    prod = _find(df.columns, ("product",))
    imp = _find(df.columns, ("impression",))
    clk = _find(df.columns, ("click",))
    conv_cols = [c for c in df.columns if "conversion" in str(c).lower()]
    spd = _find(df.columns, ("billable", "spend"), ("spend",), ("cost",))
    if not ex:
        return None
    for c in ([imp, clk, spd] + conv_cols):
        if c:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)
    df["_conv"] = sum(df[c] for c in conv_cols) if conv_cols else 0
    df["_prod"] = df[prod].astype(str) if prod else "(all)"
    has_conv = bool(conv_cols)

    # Exchange totals (for the overview table + concentration)
    g = (df.groupby(ex)
         .agg(impressions=(imp, "sum"), clicks=(clk, "sum"), spend=(spd, "sum"))
         .reset_index().rename(columns={ex: "exchange"}))
    g["ctr"] = np.where(g["impressions"] > 0, g["clicks"] / g["impressions"], 0)
    total_spend = g["spend"].sum() or 1
    g["pct_of_spend"] = g["spend"] / total_spend
    book_ctr = (g["clicks"].sum() / g["impressions"].sum()) if g["impressions"].sum() else 0

    # PRODUCT-AWARE flags: exchange x product vs that product's own CTR norm.
    agg = {"impressions": (imp, "sum"), "clicks": (clk, "sum"), "spend": (spd, "sum"),
           "conversions": ("_conv", "sum")}
    ep = (df.groupby([ex, "_prod"]).agg(**agg)
          .reset_index().rename(columns={ex: "exchange", "_prod": "product"}))
    ep["ctr"] = np.where(ep["impressions"] > 0, ep["clicks"] / ep["impressions"], 0)
    pnorm = (ep.groupby("product").apply(
        lambda x: (x["clicks"].sum() / x["impressions"].sum()) if x["impressions"].sum() else 0)
        .rename("product_ctr").reset_index())
    ep = ep.merge(pnorm, on="product", how="left")
    ep["x_over_norm"] = np.where(ep["product_ctr"] > 0, ep["ctr"] / ep["product_ctr"], 0)

    # Flag abnormally HIGH CTR for the product (invalid-traffic signal), and — only
    # for click-driven products — abnormally LOW CTR (wasted/non-engaging). CTV/Audio
    # are excluded from the low check since low clicks are expected there.
    LOW_CTR_PRODUCTS = {"Display", "Native Display", "Native Video", "Video", "Social Mirror"}
    material = (ep["spend"] >= min_spend) & (ep["impressions"] >= min_impr)
    hi_ctr = material & (ep["x_over_norm"] >= ctr_multiple)
    low_ctr = (ep["impressions"] > 30000) & (ep["product"].isin(LOW_CTR_PRODUCTS)) \
        & (ep["product_ctr"] > 0) & (ep["x_over_norm"] <= (1.0 / ctr_multiple))
    ep["_hi"] = hi_ctr
    ep["_lo"] = low_ctr & ~hi_ctr
    flags = ep[ep["_hi"] | ep["_lo"]].copy()

    def _flag_text(r):
        conv_note = (f" — BUT {int(r['conversions'])} conversions, may be worth keeping"
                     if r["conversions"] >= 5 else "")
        if r["_hi"]:
            base = (f"CTR {r['ctr']*100:.3f}% is {r['x_over_norm']:.1f}× the {r['product']} norm "
                    f"— abnormally HIGH for this ad type")
            return base + (conv_note or " — few/no conversions, likely invalid")
        base = (f"CTR {r['ctr']*100:.3f}% is {r['x_over_norm']:.2f}× the {r['product']} norm "
                f"— abnormally LOW / likely wasted")
        return base + conv_note
    if len(flags):
        flags["flag"] = flags.apply(_flag_text, axis=1)
        flags["keep_signal"] = flags["conversions"] >= 5
    flags = flags.sort_values("spend", ascending=False)

    concentration = g.sort_values("spend", ascending=False).head(1)
    top_share = float(concentration["pct_of_spend"].iloc[0]) if len(concentration) else 0

    summary = {
        "exchanges": int(len(g)),
        "book_ctr": float(book_ctr),
        "total_spend": float(g["spend"].sum()),
        "flag_count": int(len(flags)),
        "flag_spend": float(flags["spend"].sum()),
        "top_exchange": concentration["exchange"].iloc[0] if len(concentration) else None,
        "top_share": top_share,
        "has_conv": has_conv,
    }
    return {
        "summary": summary,
        "table": g.sort_values("spend", ascending=False),
        "flags": flags,
    }
