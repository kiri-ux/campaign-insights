"""
tap_adapter.py
TapClicks now exports two flat, single-sheet files (one Site Domains, one Apps)
with all fields as columns. The analysis engines expect the old six-sheet
Insights workbook, so this adapter reads the two flat files and synthesizes that
workbook in memory:
  - Site Overview  = the sites flat file (row-level, as-is)
  - App Overview   = the apps flat file (row-level, as-is)
  - Product Overview  = aggregated up from both, by BU/Client/Product
  - Strategy Overview = aggregated up from both, by BU/Client/Product/Strategy

No enrichment is invented — Product 2, Strategy Type/Name and Client Business
Unit come straight from the export. 'Internal Cost' is mapped from 'Billable
Spend' (the cost field present in the data views).
"""
import io
import os
import re
import tempfile
import numpy as np
import pandas as pd

_MEASURES = ["Impressions", "Clicks", "Post Click Conversions",
             "Post View Conversions", "Billable Spend"]

# TapClicks data-view exports use snake_case DB column names; the engines expect
# the Title Case labels from the old report. Map them (only ones present are used).
_COLMAP = {
    "date": "Date",
    "client_business_unit": "Client Business Unit",
    "client": "Client",
    "product_2": "Product 2",
    "strategy_type": "Strategy Type",
    "strategy_name": "Strategy Name",
    "campaign_id": "Campaign ID", "app_url": "App/URL", "dsp": "DSP",
    "site_app": "Site/App", "inventory_type": "Inventory Type",
    "environment": "Environment", "ad_environment": "Environment",
    "channel_type": "Channel Type",
    "site_domain": "Site Domain",
    "final_site_domain_name": "Final Site Domain Name",
    "app_name": "App Name",
    "final_app_name_use_me": "Final App Name",
    "final_app_name": "Final App Name",
    "app_id": "App ID",
    "device_type": "Device Type",
    "impressions": "Impressions",
    "clicks": "Clicks",
    "ctr": "CTR",
    "post_click_conversions": "Post Click Conversions",
    "post_view_conversions": "Post View Conversions",
    "cpm": "CPM",
    "billable_spend": "Billable Spend",
    "total_spend": "Total Spend",
}


def _canon(s):
    """Canonical form for header matching: lowercase, alphanumerics only.
    'Campaign ID', 'campaign_id', 'CampaignId ' all -> 'campaignid' — while
    'Campaign Pool ID' -> 'campaignpoolid' stays distinct."""
    import re
    return re.sub(r"[^a-z0-9]", "", str(s).lower())


# canon(header) -> engine name, built from the snake_case map AND the Title Case
# targets themselves, so any casing/spacing variant of either form matches.
_CANON_MAP = {_canon(k): v for k, v in _COLMAP.items()}
_CANON_MAP.update({_canon(v): v for v in _COLMAP.values()})


def _normalize_headers(df, canon_map=None):
    """Rename export headers to the Title Case names the engines use, matching
    on canonical form so 'Campaign Id', 'campaign_id', 'CAMPAIGN ID' all work.

    Two different headers can canonicalize to the same name — the creative data
    view ships 'Creative Name' AND 'Creative Name**', and canonical matching
    strips the punctuation that tells them apart. Renaming both to 'Creative
    Name' would leave two identically-labelled columns, and df['Creative Name']
    would then hand back a DataFrame instead of a Series, breaking every engine
    downstream. So the second and later collisions take a ' (alt)' suffix and
    are KEPT — those are the columns carrying a name when the first is blank.
    """
    cmap = canon_map or _CANON_MAP
    ren, used = {}, set()
    for c in df.columns:
        name = cmap.get(_canon(c))
        if not name:
            continue
        if name in used:
            import re as _re
            base = _re.sub(r" \(alt(?: \d+)?\)$", "", name)
            n, cand = 1, base + " (alt)"
            while cand in used:
                n += 1
                cand = f"{base} (alt {n})"
            name = cand
        used.add(name)
        ren[c] = name
    return df.rename(columns=ren) if ren else df


# The only columns the engines actually use. The export has ~30 more (CPV/CPCV,
# budgets, margins, external IDs...) that we drop on read to save a lot of memory
# on the full ~385k-row dataset.
_KEEP = ["Date", "Client Business Unit", "Client", "Product 2", "Strategy Type",
         "Strategy Name", "Campaign ID", "Site Domain", "Final Site Domain Name",
         "App Name", "Final App Name", "App ID", "Impressions", "Clicks", "CTR",
         "Post Click Conversions", "Post View Conversions", "CPM", "Billable Spend", "App/URL", "DSP", "Site/App", "Inventory Type", "Environment", "Channel Type"]
_FLOAT32 = ["CTR", "CPM"]              # display-only metrics; recomputed downstream
_CATEGORY = ["Client Business Unit", "Client", "Product 2", "Strategy Type", "Campaign ID"]


def _prune_and_downcast(df):
    keep = [c for c in _KEEP if c in df.columns]
    df = df[keep].copy()
    # money + count columns stay full precision (no overflow / cent errors on big data)
    for c in ("Impressions", "Clicks", "Post Click Conversions", "Post View Conversions"):
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)
    if "Billable Spend" in df.columns:
        df["Billable Spend"] = pd.to_numeric(df["Billable Spend"], errors="coerce").fillna(0.0)
    for c in _FLOAT32:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").astype("float32")
    # repeated text -> category is the big memory win (BU/Client/Product/Strategy)
    for c in _CATEGORY:
        if c in df.columns:
            df[c] = df[c].astype("category")
    return df


# --- creative-insights export ---------------------------------------------
# Creative-grain export (S3 prefix 'creative-insights'): one row per date x
# campaign x creative. Same dimensions as the site/app exports plus creative
# identity and the campaign pool. Kept in its OWN map/keep list so creative
# columns can never change what the site/app readers materialize.
_CREATIVE_EXTRA = {
    "creative_name": "Creative Name", "creative": "Creative Name",
    "creative_title": "Creative Name", "ad_name": "Creative Name",
    "creative_id": "Creative ID", "creative_size": "Creative Size",
    "creative_type": "Creative Type",
    # The data view carries SEVERAL creative-name fields, and which one is
    # populated varies by row: 'Creative Name' can be blank where 'Creative
    # External Name' has the real name. Each gets its own engine column so
    # resolve_creative_names() can fall back across them.
    "creative_external_name": "Creative External Name",
    "external_creative_name": "Creative External Name",
    "creative_name_external": "Creative External Name",
    "external_name": "Creative External Name",
    "creative_name_use_me": "Creative Name (alt)",
    "creative_name_2": "Creative Name (alt)",
    "creative_name_alt": "Creative Name (alt)",
    "campaign_pool_name": "Campaign Pool Name",
    "campaign_pool_id": "Campaign Pool ID",
    # Added to the data view Aug 2026. Header matching is canonical (lowercase,
    # alphanumerics only), so '25% Completed', '25_completed' and 'Completed 25%'
    # all land on the same engine column — the vendor can rename these without
    # breaking the report.
    "creative_clickthrough_url": "Clickthrough URL",
    "clickthrough_url": "Clickthrough URL",
    "click_through_url": "Clickthrough URL",
    "creative_click_through_url": "Clickthrough URL",
    "landing_page_url": "Clickthrough URL",
    "destination_url": "Clickthrough URL",
    "final_url": "Clickthrough URL",
    "preview_image_url": "Preview Image URL",
    "creative_preview_image_url": "Preview Image URL",
    "preview_url": "Preview Image URL",
    "thumbnail_url": "Preview Image URL",
    "25_completed": "25% Completed", "completed_25": "25% Completed",
    "video_25_completed": "25% Completed", "video_completions_25": "25% Completed",
    "first_quartile": "25% Completed",
    "50_completed": "50% Completed", "completed_50": "50% Completed",
    "video_50_completed": "50% Completed", "video_completions_50": "50% Completed",
    "midpoint": "50% Completed",
    "75_completed": "75% Completed", "completed_75": "75% Completed",
    "video_75_completed": "75% Completed", "video_completions_75": "75% Completed",
    "third_quartile": "75% Completed",
    "100_completed": "100% Completed", "completed_100": "100% Completed",
    "video_100_completed": "100% Completed", "video_completions_100": "100% Completed",
    "completions": "100% Completed", "video_complete": "100% Completed",
}
_CREATIVE_COLMAP = dict(_COLMAP, **_CREATIVE_EXTRA)
_CREATIVE_CANON = {_canon(k): v for k, v in _CREATIVE_COLMAP.items()}
_CREATIVE_CANON.update({_canon(v): v for v in _CREATIVE_COLMAP.values()})
_CREATIVE_KEEP = ["Date", "Client Business Unit", "Client", "Product 2",
                  "Strategy Type", "Strategy Name", "Creative Name",
                  "Creative External Name", "Creative Name (alt)",
                  "Creative Name (alt 2)", "Creative Name (alt 3)", "Creative ID",
                  "Creative Size", "Creative Type", "Campaign Pool Name",
                  "Campaign Pool ID", "Campaign ID", "Impressions", "Clicks", "CTR",
                  "Post Click Conversions", "Post View Conversions", "CPM",
                  "Billable Spend", "DSP", "Clickthrough URL", "Preview Image URL",
                  "25% Completed", "50% Completed", "75% Completed", "100% Completed",
                  "Name Source"]
_CREATIVE_QUARTILES = ["25% Completed", "50% Completed", "75% Completed", "100% Completed"]


# --- device-insights export -------------------------------------------------
# Device-grain export (S3 prefix 'device-insights', e.g.
# 'device-insights_20260822_0844_0.csv'): one row per date x campaign x device.
# Same dimensions as the creative export, with the creative identity swapped for
# the device breakdown. Own map/keep list, same as the creative reader.
_DEVICE_EXTRA = {
    "device_type": "Device Type", "device": "Device Type",
    "device_category": "Device Type", "devicetype": "Device Type",
    "operating_system": "Operating System", "os": "Operating System",
    "device_os": "Operating System",
    "browser": "Browser", "device_make": "Device Make", "make": "Device Make",
    "device_model": "Device Model", "model": "Device Model",
    "environment": "Environment", "ad_environment": "Environment",
    "channel_type": "Channel Type", "inventory_type": "Inventory Type",
    "campaign_pool_name": "Campaign Pool Name",
    "campaign_pool_id": "Campaign Pool ID",
}
_DEVICE_COLMAP = dict(_COLMAP, **_DEVICE_EXTRA)
_DEVICE_CANON = {_canon(k): v for k, v in _DEVICE_COLMAP.items()}
_DEVICE_CANON.update({_canon(v): v for v in _DEVICE_COLMAP.values()})
_DEVICE_KEEP = ["Date", "Client Business Unit", "Client", "Product 2",
                "Strategy Type", "Strategy Name", "Device Type", "Operating System",
                "Browser", "Device Make", "Device Model", "Environment",
                "Channel Type", "Inventory Type", "Campaign Pool Name",
                "Campaign Pool ID", "Campaign ID", "Impressions", "Clicks", "CTR",
                "Post Click Conversions", "Post View Conversions", "CPM",
                "Billable Spend", "DSP"]


def _prune_device(df):
    """Prune/downcast a device export. Device dimensions are low-cardinality, so
    they categorize well; the Date column stays a plain value because the
    de-duplication and reconciliation both key on it."""
    keep = [c for c in _DEVICE_KEEP if c in df.columns]
    df = df[keep].copy()
    for c in ("Impressions", "Clicks", "Post Click Conversions", "Post View Conversions"):
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)
    if "Billable Spend" in df.columns:
        df["Billable Spend"] = pd.to_numeric(df["Billable Spend"], errors="coerce").fillna(0.0)
    for c in ("CTR", "CPM"):
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").astype("float32")
    for c in ("Client Business Unit", "Product 2", "Strategy Type", "Campaign ID",
              "Campaign Pool ID", "Device Type", "Operating System", "Browser",
              "Device Make", "Environment", "Channel Type", "Inventory Type"):
        if c in df.columns:
            df[c] = df[c].astype("category")
    for c in ("Campaign ID", "Campaign Pool ID"):
        if c in df.columns and pd.api.types.is_numeric_dtype(df[c]):
            df[c] = df[c].astype("Int64").astype("string").fillna("").astype("object")
    return df


def read_device_flat(data, filename=""):
    """Read one device-insights export (xlsx or csv bytes) with normalized headers."""
    bio = io.BytesIO(data)
    if (filename or "").lower().endswith(".csv"):
        try:
            df = pd.read_csv(bio, low_memory=False,
                             usecols=lambda c: _wanted_engine_name(c, _DEVICE_CANON, _DEVICE_KEEP) is not None)
        except ValueError:
            df = None
        if df is None or not len(df.columns):
            bio.seek(0)
            df = pd.read_csv(bio, low_memory=False)
    else:
        df = _read_xlsx_slim(bio, _DEVICE_CANON, _DEVICE_KEEP)
    df = _normalize_headers(df, _DEVICE_CANON)
    if "Device Type" not in df.columns:
        return df  # let the caller report the layout problem
    return _prune_device(df)


def combine_devices(dfs, report=False):
    """Pool several device exports into one frame.

    THIS IS THE LOAD-BEARING FUNCTION for the Jul 2026 inflated-device-impressions
    issue. The vendor drops a rolling LAST-7-DAYS file every day, so any given
    delivery date appears in up to seven consecutive files. Concatenating them
    without de-duplicating multiplies that date's impressions by the number of
    files carrying it — which is exactly the shape of an inflation that hits
    impressions and clicks by the same factor.

    De-dupe keys on every dimension column (date, client, campaign, device, …)
    and keeps the LAST occurrence, so the newest file's restatement of a row
    wins. With `report=True` returns (frame, stats) where stats records how much
    duplication was removed — surfaced in the dashboard so a silent double-count
    can never look like growth again."""
    dfs = [d for d in dfs if d is not None and len(d)]
    if not dfs:
        return (pd.DataFrame(), {}) if report else pd.DataFrame()
    combined = pd.concat(dfs, ignore_index=True, sort=False)
    raw_rows = len(combined)
    raw_impr = float(pd.to_numeric(combined.get("Impressions", 0), errors="coerce").fillna(0).sum())
    dims = [c for c in combined.columns if c not in _MEASURE_COLS]
    if dims:
        combined = combined.drop_duplicates(subset=dims, keep="last")
    combined = combined.reset_index(drop=True)
    if not report:
        return combined
    kept_impr = float(pd.to_numeric(combined.get("Impressions", 0), errors="coerce").fillna(0).sum())
    stats = {"files": len(dfs), "rows_before": raw_rows, "rows_after": len(combined),
             "rows_removed": raw_rows - len(combined),
             "impressions_before": int(raw_impr), "impressions_after": int(kept_impr),
             "impressions_removed": int(raw_impr - kept_impr),
             "inflation_factor": (raw_impr / kept_impr) if kept_impr else 1.0}
    return combined, stats


def _wanted_engine_name(header, canon_map=None, keep=None):
    """Engine column name for a raw export header, or None if we don't use it."""
    cn = _canon(header)
    name = (canon_map or _CANON_MAP).get(cn)
    return name if name in (keep if keep is not None else _KEEP) else None


def _read_xlsx_slim(bio, canon_map=None, keep=None):
    """Stream the first worksheet in read_only mode, materializing ONLY the
    columns the engines use. pandas.read_excel builds every one of the export's
    ~30+ columns before we prune — on a 385k-row file that's the single biggest
    memory spike in the app, and it's what OOM-killed workers mid-pull."""
    from openpyxl import load_workbook
    wb = load_workbook(bio, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        header = next(it, None)
        if header is None:
            return pd.DataFrame()
        wanted = {}
        for i, h in enumerate(header):
            if h is None:
                continue
            name = _wanted_engine_name(h, canon_map, keep)
            if name and name not in wanted.values():
                wanted[i] = name
        if not wanted:
            # Unrecognized layout — fall back to the full read so the existing
            # error path can report it sensibly.
            bio.seek(0)
            return pd.read_excel(bio)
        data = {name: [] for name in wanted.values()}
        for row in it:
            for i, name in wanted.items():
                data[name].append(row[i] if i < len(row) else None)
        return pd.DataFrame(data)
    finally:
        wb.close()


_COMPOSITE = re.compile(r"^(.*?)\s*\(([^()]+)\)\s*$")


def _split_name_id(v):
    """'QR & Barcode Scanner: Home - Android (qrcode.barcodescanner.reader)'
    -> ('QR & Barcode Scanner: Home - Android', 'qrcode.barcodescanner.reader').
    Values without an id-looking trailing parenthetical return (v, v)."""
    v = str(v).strip()
    m = _COMPOSITE.match(v)
    if m:
        nid = m.group(2).strip()
        if "." in nid or re.fullmatch(r"\d{6,}", nid):
            return (m.group(1).strip() or nid), nid
    return v, v


def clean_app_identity(df):
    """AdLib sometimes ships composite App IDs ('Name - Android (bundle.id)').
    Split them: pure id into App ID, name part into App Name / Final App Name
    where those held the same composite (or were blank). No-op otherwise."""
    if df is None or not len(df) or "App ID" not in df.columns:
        return df
    raw = df["App ID"].astype(str)
    pairs = raw.map(_split_name_id)
    ids = pairs.map(lambda t: t[1])
    changed = ids != raw
    df = df.copy()
    df["Raw Value"] = raw  # exact source string — Reporting Zone matches on it
    if not changed.any():
        return df
    names = pairs.map(lambda t: t[0])
    for col in ("App Name", "Final App Name"):
        if col in df.columns:
            cur = df[col].astype(str)
            replace = changed & (cur.eq(raw) | cur.isin(["", "nan", "None", "NA"]))
            df[col] = cur.where(~replace, names)
    df["App ID"] = ids
    return df


def read_flat(data, filename=""):
    """Read one flat export (xlsx or csv bytes) into a DataFrame with normalized
    headers, pruned to the columns the engines use, and memory-downcast. Both
    readers load only the needed columns to keep peak memory low."""
    bio = io.BytesIO(data)
    if (filename or "").lower().endswith(".csv"):
        try:
            df = pd.read_csv(bio, low_memory=False,
                             usecols=lambda c: _wanted_engine_name(c) is not None)
        except ValueError:  # no recognizable columns — full read for the error path
            bio.seek(0)
            df = pd.read_csv(bio, low_memory=False)
    else:
        df = _read_xlsx_slim(bio)
    return clean_app_identity(_prune_and_downcast(_normalize_headers(df)))


def _prune_creative(df):
    """Prune/downcast a creative export. Deliberately does NOT touch Creative
    Name: blanks are the thing we're hunting, so they stay exactly as the vendor
    sent them (NaN stays NaN, whitespace stays whitespace) and the column is
    never categorized or filled."""
    keep = [c for c in _CREATIVE_KEEP if c in df.columns]
    df = df[keep].copy()
    for c in ["Impressions", "Clicks", "Post Click Conversions",
              "Post View Conversions"] + _CREATIVE_QUARTILES:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)
    # URLs stay verbatim strings: a blank Preview Image URL is a finding, and the
    # clickthrough URL's query string is what the UTM analysis reads.
    for c in ("Clickthrough URL", "Preview Image URL"):
        if c in df.columns:
            df[c] = df[c].astype("object")
    # IDs are labels, not measures. A column with any blank reads as float64, so
    # '1001801' would render (and export) as '1001801.0' — nobody can paste that
    # into the DSP. Force them to clean strings, blanks stay blank.
    for c in ("Creative ID", "Campaign ID", "Campaign Pool ID"):
        if c in df.columns and pd.api.types.is_numeric_dtype(df[c]):
            df[c] = (df[c].astype("Int64").astype("string")
                     .fillna("").astype("object"))
    if "Billable Spend" in df.columns:
        df["Billable Spend"] = pd.to_numeric(df["Billable Spend"], errors="coerce").fillna(0.0)
    for c in ("CTR", "CPM"):
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").astype("float32")
    for c in ("Client Business Unit", "Product 2", "Strategy Type", "Campaign ID",
              "Campaign Pool ID", "Creative Type", "Creative Size"):
        if c in df.columns:
            df[c] = df[c].astype("category")
    return df


# Fallback order for a creative's name. The vendor's files carry a name on 100%
# of rows; the blanks are created by the data view, and which of its name fields
# is populated varies row to row — so a blank in the first column is not a
# missing name until every one of these is blank too.
_NAME_FALLBACKS = ["Creative Name", "Creative External Name", "Creative Name (alt)",
                   "Creative Name (alt 2)", "Creative Name (alt 3)"]
_NAME_SOURCE_COL = "Name Source"


def _blank_mask(s):
    """True where a text column is empty in any of the ways an export manages."""
    t = s.astype(str).str.strip().str.lower()
    return s.isna() | t.isin(["", "nan", "none", "null", "n/a", "na", "-", "--",
                              "#n/a", "0", "."])


def resolve_creative_names(df, report=False):
    """Fill a blank Creative Name from the export's other name fields, then from
    the same Creative ID elsewhere in the data. Records where each name came from.

    Two independent recoveries, in order of trust:

    1. **Sibling columns** — 'Creative External Name' / 'Creative Name (alt)'.
       Same row, same vendor, so this is a straight read, not a guess.
    2. **The same Creative ID on another row** — one ID is one creative, so a
       name found on any row of that ID applies to all of them. The most
       frequent non-blank spelling wins. Only runs where an ID exists.

    Anything still blank afterwards is a genuine missing name and stays blank —
    the point of the report is to catch those, so nothing is invented here.
    """
    stats = {"rows": int(len(df)), "blank_before": 0, "recovered": 0,
             "blank_after": 0, "by_source": {}, "recovered_impressions": 0}
    if not len(df) or "Creative Name" not in df.columns:
        return (df, stats) if report else df

    df = df.copy()
    name = df["Creative Name"].astype("object")
    blank = _blank_mask(name)
    source = pd.Series(np.where(blank, "", "vendor"), index=df.index, dtype="object")
    # Second pass over pooled files: keep the provenance the first pass recorded,
    # or every already-recovered row would be relabelled 'vendor' and the report
    # would understate how much of the library the vendor actually named.
    if _NAME_SOURCE_COL in df.columns:
        prior = df[_NAME_SOURCE_COL].astype("object")
        keep = prior.notna() & (prior != "") & (prior != "(still blank)") & ~blank
        source = source.where(~keep, prior)

    for col in _NAME_FALLBACKS[1:]:
        if col not in df.columns or not blank.any():
            continue
        alt = df[col].astype("object")
        take = blank & ~_blank_mask(alt)
        if take.any():
            name = name.where(~take, alt)
            source = source.where(~take, col)
            blank = _blank_mask(name)

    if blank.any() and "Creative ID" in df.columns:
        cid = df["Creative ID"].astype(str).str.strip()
        known = pd.DataFrame({"_id": cid[~blank], "_n": name[~blank].astype(str)})
        known = known[(known["_id"] != "") & (known["_id"].str.lower() != "nan")]
        if len(known):
            # most frequent spelling per ID — vendor exports restate names
            best = (known.groupby("_id")["_n"].agg(
                lambda s: s.value_counts().idxmax()))
            filled = cid.map(best)
            take = blank & filled.notna()
            if take.any():
                name = name.where(~take, filled)
                source = source.where(~take, "matched on Creative ID")
                blank = _blank_mask(name)

    df["Creative Name"] = name
    src = source.where(source != "", "(still blank)")
    df[_NAME_SOURCE_COL] = src
    # Stats describe the FINAL state of the frame, not just this pass, so they
    # stay true whether resolution ran once (single upload) or twice (pooled).
    rec = src.isin(_NAME_FALLBACKS[1:] + ["matched on Creative ID"])
    stats["recovered"] = int(rec.sum())
    stats["blank_after"] = int((src == "(still blank)").sum())
    stats["blank_before"] = stats["recovered"] + stats["blank_after"]
    stats["by_source"] = {k: int(v) for k, v in src[rec].value_counts().items()}
    if "Impressions" in df.columns:
        stats["recovered_impressions"] = int(
            pd.to_numeric(df.loc[rec, "Impressions"], errors="coerce").fillna(0).sum())
    return (df, stats) if report else df


def read_creative_flat(data, filename=""):
    """Read one creative-insights export (xlsx or csv bytes) into a DataFrame
    with normalized headers. Blank creative names are preserved verbatim."""
    bio = io.BytesIO(data)
    if (filename or "").lower().endswith(".csv"):
        try:
            df = pd.read_csv(
                bio, low_memory=False, keep_default_na=True,
                usecols=lambda c: _wanted_engine_name(c, _CREATIVE_CANON, _CREATIVE_KEEP) is not None)
        except ValueError:
            df = None
        if df is None or not len(df.columns):
            # Unrecognized layout — a column-filtered read of a file we don't
            # know yields an EMPTY frame, which downstream reads as "no data"
            # instead of "wrong file". Re-read in full so the caller can say so.
            bio.seek(0)
            df = pd.read_csv(bio, low_memory=False)
    else:
        df = _read_xlsx_slim(bio, _CREATIVE_CANON, _CREATIVE_KEEP)
    df = _normalize_headers(df, _CREATIVE_CANON)
    if "Creative Name" not in df.columns:
        return df  # let the caller report the layout problem
    # Resolve names BEFORE pruning, while the sibling name columns are still here.
    return _prune_creative(resolve_creative_names(df))


def combine_creatives(dfs, report=False):
    """Concat several creative exports, de-duped on the dimension columns
    (rolling windows restate rows) keeping the newest file's version.

    Name resolution runs again after the concat: a creative that is blank in
    every row of today's file may be named in yesterday's, and the ID match can
    only see that once the files are pooled.
    """
    dfs = [d for d in dfs if d is not None and len(d)]
    if not dfs:
        return (pd.DataFrame(), {}) if report else pd.DataFrame()
    combined = pd.concat(dfs, ignore_index=True, sort=False)
    dims = [c for c in combined.columns if c not in _MEASURE_COLS]
    if dims:
        # dropna=False semantics: blank creative names must survive de-dupe
        combined = combined.drop_duplicates(subset=dims, keep="last")
    combined = combined.reset_index(drop=True)
    out = resolve_creative_names(combined, report=report)
    return out if report else out


_MEASURE_COLS = set(_MEASURES) | {"CTR", "CPM", "Total Spend"}


def combine_flats(dfs):
    """Concat several date-ranged flat exports into one frame. Overlapping exports
    (rolling windows / restated data) produce duplicate dimension rows — keep the
    LAST occurrence (files are processed oldest-first, so the newest export's
    version of any restated row wins) and drop the rest so pooled metrics never
    double-count."""
    dfs = [d for d in dfs if d is not None and len(d)]
    if not dfs:
        return pd.DataFrame()
    combined = pd.concat(dfs, ignore_index=True, sort=False)
    dims = [c for c in combined.columns if c not in _MEASURE_COLS]
    if dims:
        combined = combined.drop_duplicates(subset=dims, keep="last")
    return combined.reset_index(drop=True)


# Real TLDs a placement domain plausibly ends with; used to split the TTD/DV360
# combined export's single App/URL column into site rows vs app rows.
_TLDS = {"com","net","org","co","io","tv","us","uk","ca","de","fr","es","it","nl",
         "se","jp","au","in","br","mx","fm","gg","me","app","edu","gov","info","biz","news"}
_REV_DNS = {"com","net","org","co","io","tv","jp","de","uk","us","air","me","app","tw","kr"}


def _looks_like_app(v):
    v = str(v).strip().lower()
    if not v or v in ("nan", "none"):
        return True
    if re.fullmatch(r"(id)?\d{6,}", v):
        return True                                   # iOS store ids
    if "." not in v:
        return True                                   # bare app names
    labels = v.split(".")
    if len(labels) >= 3 and labels[0] in _REV_DNS:
        return True                                   # reverse-DNS bundle ids
    return labels[-1] not in _TLDS                    # real TLD -> site




# App inventory that masquerades as a domain (e.g. the AOL Desktop App serves
# under cdn.desktop.aol.com). Extend via TTDDV_FORCE_APP / TTDDV_FORCE_SITE
# (comma-separated, matched case-insensitively on the id).
_KNOWN_APP_DOMAINS = {"cdn.desktop.aol.com"}
_TYPE_COLS = ("Site/App", "Inventory Type", "Environment", "Channel Type")


def _forced(name_set_env, builtin=frozenset()):
    vals = {v.strip().lower() for v in os.environ.get(name_set_env, "").split(",") if v.strip()}
    return vals | set(builtin)


def _classify_ttddv_rows(df, ids):
    """Per-row app/site decision: explicit type column wins, then the forced
    override lists, then shape heuristics."""
    forced_app = _forced("TTDDV_FORCE_APP", _KNOWN_APP_DOMAINS)
    forced_site = _forced("TTDDV_FORCE_SITE")
    ids_l = ids.astype(str).str.lower()
    isapp = ids_l.map(_looks_like_app)          # shape baseline
    tcol = next((c for c in _TYPE_COLS if c in df.columns), None)
    if tcol:
        t = df[tcol].astype(str).str.lower()
        explicit_app = t.str.contains("app", na=False)
        explicit_site = t.str.contains("site|web", regex=True, na=False) & ~explicit_app
        isapp = isapp.where(~(explicit_app | explicit_site), explicit_app)
    isapp = isapp.where(~ids_l.isin(forced_site), False)
    isapp = isapp.where(~ids_l.isin(forced_app), True)
    return isapp


def split_ttddv(df, dsp_label="TTD/DV360"):
    """Split the TTD/DV360 combined export (single App/URL column) into
    site-shaped and app-shaped frames tagged with a DSP column, so they merge
    straight into the regular site/app analysis. Composite 'Name (bundle.id)'
    values are parsed so App Name and App ID land in their own columns."""
    empty = pd.DataFrame()
    if df is None or not len(df) or "App/URL" not in df.columns:
        return empty, empty
    df = df.copy()
    df["DSP"] = dsp_label
    pairs = df["App/URL"].map(_split_name_id)
    names = pairs.map(lambda t: t[0])
    ids = pairs.map(lambda t: t[1])
    isapp = _classify_ttddv_rows(df, ids)
    sites = df[~isapp].copy()
    apps = df[isapp].copy()
    if len(sites):
        sites["Raw Value"] = sites["App/URL"].astype(str)
        sites["Site Domain"] = ids[~isapp]
        sites["Final Site Domain Name"] = ids[~isapp]
        sites = sites.drop(columns=["App/URL", *[c for c in _TYPE_COLS if c in sites.columns]])
    if len(apps):
        apps["Raw Value"] = apps["App/URL"].astype(str)
        apps["App Name"] = names[isapp]
        apps["Final App Name"] = names[isapp]
        apps["App ID"] = ids[isapp]
        apps = apps.drop(columns=["App/URL", *[c for c in _TYPE_COLS if c in apps.columns]])
    return sites, apps


def filter_date_range(df, start_iso, end_iso):
    """Keep only rows whose Date falls within [start, end] inclusive. Frames
    without a Date column pass through untouched (nothing to filter on)."""
    if df is None or not len(df) or "Date" not in df.columns:
        return df
    d = pd.to_datetime(df["Date"], errors="coerce")
    import datetime as _dt
    start = pd.Timestamp(_dt.date.fromisoformat(start_iso))
    end = pd.Timestamp(_dt.date.fromisoformat(end_iso))
    return df[(d >= start) & (d <= end)].reset_index(drop=True)


def _bu_col(df):
    for c in ("Business Unit", "Client Business Unit"):
        if c in df.columns:
            return c
    return df.columns[0]


def _overview(df, keys):
    """Aggregate row-level delivery up to an overview grain, renaming the flat
    measure columns to the Overview names the insights engine expects."""
    d = df.copy()
    for m in _MEASURES:
        if m in d.columns:
            d[m] = pd.to_numeric(d[m], errors="coerce").fillna(0)
    agg = {"Impressions": ("Impressions", "sum"), "Clicks": ("Clicks", "sum")}
    if "Post Click Conversions" in d.columns:
        agg["Click Conversions"] = ("Post Click Conversions", "sum")
    if "Post View Conversions" in d.columns:
        agg["View-throughs"] = ("Post View Conversions", "sum")
    if "Billable Spend" in d.columns:
        agg["Internal Cost"] = ("Billable Spend", "sum")
    g = d.groupby(keys, dropna=False, observed=True).agg(**agg).reset_index()
    # groupby on category keys yields category columns; the small overview frames
    # are tiny, so cast keys back to str to avoid category edge cases downstream.
    for k in keys:
        if k in g.columns and str(g[k].dtype) == "category":
            g[k] = g[k].astype(str)
    # guarantee the columns downstream aggregations reference
    for col in ("Click Conversions", "View-throughs", "Internal Cost"):
        if col not in g.columns:
            g[col] = 0
    return g


def build_frames(sites_df, apps_df):
    """Build the four analysis frames (site/app/product/strategy) directly, with
    NO xlsx round-trip — keeps the memory downcasting and avoids a duplicate copy."""
    combined = pd.concat([sites_df, apps_df], ignore_index=True, sort=False)
    bu = _bu_col(combined)
    prod_keys = [k for k in (bu, "Client", "Product 2") if k in combined.columns]
    prod = _overview(combined, prod_keys).rename(columns={"Product 2": "Product"})
    strat_keys = [k for k in (bu, "Client", "Product 2", "Strategy Type", "Strategy Name")
                  if k in combined.columns]
    strat = _overview(combined, strat_keys).rename(columns={"Product 2": "Product"})
    del combined
    import gc
    gc.collect()
    return {"site": sites_df, "app": apps_df, "product": prod, "strategy": strat}


def synthesize_workbook(sites_df, apps_df):
    """Write the four frames to a temp .xlsx and return its path (manual-upload
    path; the automated pull uses build_frames to skip the xlsx round-trip)."""
    f = build_frames(sites_df, apps_df)
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp.close()
    with pd.ExcelWriter(tmp.name, engine="openpyxl") as xl:
        f["site"].to_excel(xl, sheet_name="Site Overview", index=False)
        f["app"].to_excel(xl, sheet_name="App Overview", index=False)
        f["product"].to_excel(xl, sheet_name="Product Overview", index=False)
        f["strategy"].to_excel(xl, sheet_name="Strategy Overview", index=False)
    return tmp.name
